#!/usr/bin/env python3
"""Render a styled .xlsx workbook from a JSON spec, deterministically, via openpyxl.

The determinism doctrine says the LLM writes only the data and structure, and a
script renders the artifact the same way on every run. This script is that
renderer: it reads a JSON spec describing sheets, columns, rows, and per-column
number formats, then writes a workbook with a fixed design — one reused header
style, a frozen header row, content-derived column widths, and the spec's number
formats. No styling is invented per cell beyond what the spec drives.

    render.py build <spec.json> <out.xlsx>     # render a spec into a styled .xlsx
    render.py --selftest                       # build + reopen an in-code fixture

Spec shape (one object; ``sheets`` is a non-empty list):

    {
      "sheets": [
        {
          "name": "Sheet1",
          "columns": ["Region", "Q1", "Q2"],
          "rows": [["West", 10, 20], ["East", 5, 8]],
          "number_formats": {"Q1": "#,##0", "Q2": "#,##0"}
        }
      ]
    }

Each sheet needs a ``name`` and a non-empty ``columns`` list. Each row's length
must not exceed its sheet's column count. ``rows`` and ``number_formats`` are
optional; an unknown column in ``number_formats`` is rejected at the boundary.
"""

from __future__ import annotations

import argparse
import json
import sys
import tempfile
from pathlib import Path
from typing import Any

# --- design tokens, defined once and reused for every header cell ---
_HEADER_FONT_COLOR = "FFFFFF"
_HEADER_FILL_COLOR = "1F3A5F"  # deep slate
_HEADER_FONT_NAME = "Calibri"
_HEADER_FONT_SIZE = 11
_MIN_COL_WIDTH = 8
_MAX_COL_WIDTH = 60
_WIDTH_PADDING = 2
_FREEZE = "A2"  # row 1 stays on screen while scrolling


def _load_openpyxl() -> Any:
    """Return the openpyxl module, or raise a clear error when it is absent."""
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - exercised only without openpyxl
        raise RuntimeError(
            "openpyxl is required to render .xlsx files; install it with "
            "'pip install openpyxl' (or 'uv run --with openpyxl ...')."
        ) from exc
    return openpyxl


def _validate_spec(spec: Any) -> list[dict]:
    """Validate the spec at the boundary and return its sheet list, or raise ValueError."""
    if not isinstance(spec, dict):
        raise ValueError("spec must be a JSON object")
    sheets = spec.get("sheets")
    if not isinstance(sheets, list) or not sheets:
        raise ValueError("spec must have a non-empty 'sheets' list")
    for index, sheet in enumerate(sheets):
        _validate_sheet(sheet, index)
    return sheets


def _validate_sheet(sheet: Any, index: int) -> None:
    """Validate one sheet entry at position ``index``, or raise ValueError on a violation."""
    where = f"sheet[{index}]"
    if not isinstance(sheet, dict):
        raise ValueError(f"{where} must be an object")
    name = sheet.get("name")
    if not isinstance(name, str) or name.strip() == "":
        raise ValueError(f"{where} needs a non-empty string 'name'")
    columns = sheet.get("columns")
    if not isinstance(columns, list) or not columns:
        raise ValueError(f"{where} ('{name}') needs a non-empty 'columns' list")
    if not all(isinstance(col, str) for col in columns):
        raise ValueError(f"{where} ('{name}') columns must all be strings")
    _validate_rows(sheet.get("rows", []), columns, name, where)
    _validate_formats(sheet.get("number_formats", {}), columns, name, where)


def _validate_rows(rows: Any, columns: list, name: str, where: str) -> None:
    """Reject a non-list ``rows`` or any row wider than the column count."""
    if not isinstance(rows, list):
        raise ValueError(f"{where} ('{name}') 'rows' must be a list")
    width = len(columns)
    for row_index, row in enumerate(rows):
        if not isinstance(row, list):
            raise ValueError(f"{where} ('{name}') row[{row_index}] must be a list")
        if len(row) > width:
            raise ValueError(
                f"{where} ('{name}') row[{row_index}] has {len(row)} cells, "
                f"exceeding the {width} column(s)"
            )


def _validate_formats(formats: Any, columns: list, name: str, where: str) -> None:
    """Reject a non-mapping ``number_formats`` or a key naming an unknown column."""
    if not isinstance(formats, dict):
        raise ValueError(f"{where} ('{name}') 'number_formats' must be an object")
    known = set(columns)
    for key in formats:
        if key not in known:
            raise ValueError(f"{where} ('{name}') number_formats names unknown column '{key}'")


def _column_widths(columns: list[str], rows: list[list]) -> list[float]:
    """Return one clamped width per column from the longest content in that column."""
    widths: list[float] = []
    for col_index, header in enumerate(columns):
        longest = len(str(header))
        for row in rows:
            if col_index < len(row) and row[col_index] is not None:
                longest = max(longest, len(str(row[col_index])))
        clamped = max(_MIN_COL_WIDTH, min(_MAX_COL_WIDTH, longest + _WIDTH_PADDING))
        widths.append(float(clamped))
    return widths


def _write_sheet(openpyxl: Any, ws: Any, sheet: dict) -> None:
    """Write one validated sheet onto worksheet ``ws`` with the fixed design."""
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_font = Font(
        name=_HEADER_FONT_NAME,
        size=_HEADER_FONT_SIZE,
        bold=True,
        color=_HEADER_FONT_COLOR,
    )
    header_fill = PatternFill("solid", fgColor=_HEADER_FILL_COLOR)

    columns: list[str] = sheet["columns"]
    rows: list[list] = sheet.get("rows", [])
    formats: dict[str, str] = sheet.get("number_formats", {})

    # Header band: one shared font and fill, applied once per column.
    for col_index, label in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_index, value=label)
        cell.font = header_font
        cell.fill = header_fill

    # Data rows: values from the spec, with per-column number formats applied.
    for row_offset, row in enumerate(rows):
        excel_row = row_offset + 2
        for col_index, value in enumerate(row, start=1):
            cell = ws.cell(row=excel_row, column=col_index, value=value)
            column_name = columns[col_index - 1]
            fmt = formats.get(column_name)
            if fmt is not None:
                cell.number_format = fmt

    # Frame: content-derived widths and a frozen header row.
    for col_index, width in enumerate(_column_widths(columns, rows), start=1):
        ws.column_dimensions[get_column_letter(col_index)].width = width
    ws.freeze_panes = _FREEZE


def build(spec_path: Path, out_path: Path) -> int:
    """Read a JSON spec, render the styled workbook, and write it to ``out_path``."""
    openpyxl = _load_openpyxl()
    if not spec_path.is_file():
        print(f"error: spec not found: {spec_path}", file=sys.stderr)
        return 2
    try:
        spec = json.loads(spec_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        print(f"error: spec is not valid JSON: {exc}", file=sys.stderr)
        return 2
    try:
        sheets = _validate_spec(spec)
    except ValueError as exc:
        print(f"error: invalid spec: {exc}", file=sys.stderr)
        return 2

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # drop the implicit first sheet; the spec owns every sheet
    for sheet in sheets:
        ws = wb.create_sheet(title=sheet["name"])
        _write_sheet(openpyxl, ws, sheet)
    wb.save(str(out_path))
    print(f"wrote workbook: {out_path} ({len(sheets)} sheet(s))")
    return 0


def selftest() -> int:
    """Build an in-code fixture into a temp dir, reopen it, and assert the contents."""
    try:
        openpyxl = _load_openpyxl()
    except RuntimeError:
        print("xlsx render selftest: skipped (openpyxl absent)")
        return 0

    fixture = {
        "sheets": [
            {
                "name": "Sales",
                "columns": ["Region", "Q1", "Q2"],
                "rows": [["West", 10, 20], ["East", 5, 8]],
                "number_formats": {"Q1": "#,##0", "Q2": "#,##0"},
            }
        ]
    }

    # Boundary validation rejects a row wider than the columns.
    try:
        _validate_spec({"sheets": [{"name": "X", "columns": ["A"], "rows": [[1, 2]]}]})
    except ValueError:
        pass
    else:
        print("xlsx render selftest: FAIL (overlong row accepted)", file=sys.stderr)
        return 1

    with tempfile.TemporaryDirectory(prefix="xlsx-render-selftest.") as tmp:
        spec_path = Path(tmp) / "spec.json"
        out_path = Path(tmp) / "out.xlsx"
        spec_path.write_text(json.dumps(fixture), encoding="utf-8")

        if build(spec_path, out_path) != 0:
            print("xlsx render selftest: FAIL (build did not exit zero)", file=sys.stderr)
            return 1
        if not out_path.is_file():
            print("xlsx render selftest: FAIL (no workbook written)", file=sys.stderr)
            return 1

        wb = openpyxl.load_workbook(str(out_path))
        ws = wb["Sales"]
        if ws["A1"].value != "Region":
            print(
                f"xlsx render selftest: FAIL (header A1 was {ws['A1'].value!r})",
                file=sys.stderr,
            )
            return 1
        if ws["B2"].value != 10:
            print(
                f"xlsx render selftest: FAIL (data B2 was {ws['B2'].value!r})",
                file=sys.stderr,
            )
            return 1
        if ws.freeze_panes != _FREEZE:
            print(
                f"xlsx render selftest: FAIL (freeze_panes was {ws.freeze_panes!r})",
                file=sys.stderr,
            )
            return 1
        if ws["B2"].number_format != "#,##0":
            print(
                f"xlsx render selftest: FAIL (B2 number_format was {ws['B2'].number_format!r})",
                file=sys.stderr,
            )
            return 1

    print("xlsx render selftest: ok")
    return 0


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--selftest", action="store_true", help="Run the self-test and exit.")
    sub = parser.add_subparsers(dest="command")
    build_parser = sub.add_parser("build", help="Render a JSON spec into a styled .xlsx.")
    build_parser.add_argument("spec", help="Path to the JSON spec.")
    build_parser.add_argument("out", help="Output path for the .xlsx workbook.")
    args = parser.parse_args(argv)

    if args.selftest:
        return selftest()
    if args.command == "build":
        return build(Path(args.spec), Path(args.out))
    parser.error("choose a command: 'build <spec.json> <out.xlsx>' or '--selftest'")
    return 2  # pragma: no cover - argparse exits first


if __name__ == "__main__":
    raise SystemExit(main())
